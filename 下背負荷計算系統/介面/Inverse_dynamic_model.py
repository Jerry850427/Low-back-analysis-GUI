# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 17:13:31 2020

@author: 高漢佑
"""
from openpyxl import load_workbook, Workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
from openpyxl.styles import Font, Color, Fill, PatternFill
from openpyxl.utils import get_column_letter
import statistics 
import os


seg = []
bottom_up_f = []
bottom_up_m = []
bottom_up_c = []
top_down_f = []
top_down_m = []
top_down_c = []
top_down_f2 = []
top_down_m2 = []
top_down_c2 = []
top_down_f3 = []
top_down_m3 = []
top_down_c3 = []
title_list1 = ['Fx(Bottom-up)','Fy(Bottom-up)','Fz(Bottom-up)','Mx(Bottom-up)','My(Bottom-up)','Mz(Bottom-up)','M_total(Bottom-up)','CF(Bottom-up)']              
title_list2 = ['Fx(Top-down1)','Fy(Top-down1)','Fz(Top-down1)','Mx(Top-down1)','My(Top-down1)','Mz(Top-down1)','M_total(Top-down1)','CF(Top-down1)']
title_list3 = ['Fx(Top-down2)','Fy(Top-down2)','Fz(Top-down2)','Mx(Top-down2)','My(Top-down2)','Mz(Top-down2)','M_total(Top-down2)','CF(Top-down2)']
title_list4 = ['Fx(Top-down3)','Fy(Top-down3)','Fz(Top-down3)','Mx(Top-down3)','My(Top-down3)','Mz(Top-down3)','M_total(Top-down3)','CF(Top-down3)']
g = np.array([0, 0, -9.8])
def segment_properties(gender):         #get segment propertities based on gender
    if gender == "男性":
        f = open("properties/Male.txt",'r') 
    elif gender == "女性":
        f = open("properties/Female.txt",'r')
    else:
        sys.exit("Error: invalid gender number!")
    return f

def seg_ma(ha, mass, box_mass, seg_acc, seg):                           #segment mass * segment linear acceleration    
    ha[:,:3] = (box_mass/2)*seg_acc[:,27:30]
    ha[:,3:6] = (box_mass/2)*seg_acc[:,30:33]
    for i in range(0,seg_acc.shape[1],3):        
        seg_acc[:,i:i+3] *= mass*seg[int(i/3)][1]     
    
def seg_moment_inertia(seg_ang, inertia):               #segment inertia * angular acceleration
    for i in range(0,seg_ang.shape[1],3):
        seg_ang[:,i:i+3] *= inertia[:,i:i+3]
        
def bottom_up(mass, frame, FP_data, l5s1_pos, seg_acc, seg_com, seg_ang):                        #bottom_up model (Hof, 1992)
    k = mass*g  
    seg_mg = np.zeros(shape=(1,3))
    v2 = np.zeros(shape=(frame-4,3))
    v3 = np.zeros(shape=(frame-4,3))
    v4 = np.zeros(shape=(frame-4,3))
    lower_seg_acc = np.zeros(shape=(frame-4,3))
    for i in range(7):
        seg_mg += seg[i][1]*mass*g    
    
    FP_data[:,:3] -= l5s1_pos           #(Lcop-L5S1)
    FP_data[:,9:12] -= l5s1_pos         #(Rcop-L5S1)
    v1 = np.cross(FP_data[:,:3],FP_data[:,3:6]) + np.cross(FP_data[:,9:12],FP_data[:,12:15])  #(cop-l5s1)xGRF
    
    for i in range(0,21,3):
        lower_seg_acc = lower_seg_acc+seg_acc[:,i:i+3]
        v2 = v2 + np.cross(seg_com[:,i:i+3],k*seg[int(i/3)][1])
        v3 = v3 + np.cross(seg_com[:,i:i+3],seg_acc[:,i:i+3])
        v4 += seg_ang[:,i:i+3]
    
    bottom_up_L5S1_force = -1*FP_data[:,3:6]-FP_data[:,12:15]-seg_mg+lower_seg_acc #L5S1 3D force = -GRF-lower_seg_mg+lower_seg_acc                
    bottom_up_L5S1_moment = -v1-v2+v3+v4-FP_data[:,6:9]-FP_data[:,15:18]           #L5S1 3D moment = -(cop-l5s1)xGRF-(com-l5s1)xmg+(com-l5s1)xma+Ia-free_moment
    
    return bottom_up_L5S1_force,  bottom_up_L5S1_moment     

def top_down(mass, frame, seg_acc, FP_data, box_pos, l5s1_pos, seg_com, seg_ang):  #top_down model (Hof, 1992)
    k = mass*g
    upper_seg_mg = np.zeros(shape=(1,3))
    cross_mg = np.zeros(shape=(frame-4,3))
    cross_ma = np.zeros(shape=(frame-4,3))
    upper_seg_acc = np.zeros(shape=(frame-4,3))
    total_seg_acc = np.zeros(shape=(frame-4,3))
    ia = np.zeros(shape=(frame-4,3))
    for i in range(0,seg_acc.shape[1],3):
        total_seg_acc += seg_acc[:,i:i+3]  
          
    hand_force = -1*(FP_data[:,3:6]+FP_data[:,12:15]+mass*g-total_seg_acc)      #hand force (Gert.S.Faber et.al(2013))
    hand_moment = np.cross((box_pos-l5s1_pos),hand_force)                       #hand moment = (box COM - L5S1)x hand force
    
    for i in range(7,15,1):
        upper_seg_mg +=  seg[i][1]*mass*g 
    for i in range(21,seg_acc.shape[1],3):
        upper_seg_acc += seg_acc[:,i:i+3]
        cross_mg = cross_mg + np.cross(seg_com[:,i:i+3],k*seg[int(i/3)][1])
        cross_ma = cross_ma + np.cross(seg_com[:,i:i+3],seg_acc[:,i:i+3])
        ia += seg_ang[:,i:i+3]     
        
    top_down1_L5S1_force = -1*(-hand_force-upper_seg_mg+upper_seg_acc)          #L5S1 3D force = -hand force-upper_seg_mg+upper_seg_acc
    top_down1_L5S1_moment = -1*(-1*hand_moment-cross_mg+cross_ma+ia)            #L5S1 3D moment = hand moment+(com-l5s1)xmg-(com-l5s1)xma-Ia
    return top_down1_L5S1_force, top_down1_L5S1_moment    

def top_dwon2(mass, frame, box_mass, seg_acc2, ha, seg_com, seg_ang):           #top_down model (Hof, 1992), box weight add to hand
    k = mass*g    
    upper_seg_mg = np.zeros(shape=(1,3))
    cross_mg = np.zeros(shape=(frame-4,3))
    cross_ma = np.zeros(shape=(frame-4,3))
    upper_seg_acc = np.zeros(shape=(frame-4,3))    
    ia = np.zeros(shape=(frame-4,3))
    for i in range(7,15,1):
        upper_seg_mg +=  seg[i][1]*mass*g
    upper_seg_mg +=  box_mass*g
    seg_acc2[:,27:33] += ha 
    
    for i in range(21,seg_acc2.shape[1],3):        
        upper_seg_acc += seg_acc2[:,i:i+3]        
        cross_ma = cross_ma + np.cross(seg_com[:,i:i+3],seg_acc2[:,i:i+3])
        ia += seg_ang[:,i:i+3]
        if i == 27 or i == 30:            
            cross_mg = cross_mg + np.cross(seg_com[:,i:i+3],g*(mass*seg[int(i/3)][1]+box_mass/2))
        else:
            cross_mg = cross_mg + np.cross(seg_com[:,i:i+3],k*seg[int(i/3)][1])
        
    top_down2_L5S1_force = -1*(-upper_seg_mg+upper_seg_acc)                     
    top_down2_L5S1_moment = -1*(-cross_mg+cross_ma+ia)
    return top_down2_L5S1_force, top_down2_L5S1_moment

def top_down3(mass, frame, seg_acc, FP_data, box_pos, l5s1_pos, seg_com, seg_ang, box_acc, box_mass):  #top_down model (Hof, 1992)
    k = mass*g
    upper_seg_mg = np.zeros(shape=(1,3))
    cross_mg = np.zeros(shape=(frame-4,3))
    cross_ma = np.zeros(shape=(frame-4,3))
    upper_seg_acc = np.zeros(shape=(frame-4,3))
    total_seg_acc = np.zeros(shape=(frame-4,3))
    ia = np.zeros(shape=(frame-4,3))
    for i in range(0,seg_acc.shape[1],3):
        total_seg_acc += seg_acc[:,i:i+3]  
          
    hand_force = -1*(box_mass*box_acc-box_mass*g)                               #hand force(box*(acc-g)) (Gert.S.Faber et.al(2018))
    hand_moment = np.cross((box_pos-l5s1_pos),hand_force)                       #hand moment = (box COM - L5S1)x hand force
    
    for i in range(7,15,1):
        upper_seg_mg +=  seg[i][1]*mass*g 
    for i in range(21,seg_acc.shape[1],3):
        upper_seg_acc += seg_acc[:,i:i+3]
        cross_mg = cross_mg + np.cross(seg_com[:,i:i+3],k*seg[int(i/3)][1])
        cross_ma = cross_ma + np.cross(seg_com[:,i:i+3],seg_acc[:,i:i+3])
        ia += seg_ang[:,i:i+3]     
        
    top_down3_L5S1_force = -1*(-hand_force-upper_seg_mg+upper_seg_acc)          #L5S1 3D force = -hand force-upper_seg_mg+upper_seg_acc
    top_down3_L5S1_moment = -1*(-1*hand_moment-cross_mg+cross_ma+ia)            #L5S1 3D moment = hand moment+(com-l5s1)xmg-(com-l5s1)xma-Ia
    return top_down3_L5S1_force, top_down3_L5S1_moment

def compressive_force(frame, moment):
    c = np.zeros(shape=(frame-4,1)) 
    
    for i,j in zip(moment,range(c.shape[0])):        
        cf = 1067.6+1.219*i[0]+0.083*(i[0]**2)-0.0001*(i[0]**3)+3.229*i[1]+0.119*(i[1]**2)-0.0001*(i[1]**3)+0.862*i[2]+0.393*(i[2]**2)-0.0001*(i[2]**3)
        c[j,0] = cf
    return c    

def get_data(motion_dict, file):
    data_file = file
    box_mass = float(motion_dict["box"])
    mass = float(motion_dict["weight"])
    k = segment_properties(motion_dict["sex"])
    for line in k:
        (key, val) = line.split()
        val = float(val)
        seg.append([key,val])
    parameters = pd.read_excel(data_file,sheet_name = 'Parameters')
    force_data = pd.read_excel(data_file,sheet_name = 'Force_Data')
    inertia = pd.read_excel(data_file,sheet_name = 'Inertia')
    joint = pd.read_excel(data_file,sheet_name = 'Landmark')
    frame = parameters.shape[0]                            #total number of frame
    FP_data = force_data.iloc[2:frame-2,1:].to_numpy()
    seg_acc = parameters.iloc[2:frame-2,1:46].to_numpy()
    seg_ang = parameters.iloc[2:frame-2,46:91].to_numpy()
    seg_com = parameters.iloc[2:frame-2,91:136].to_numpy()
    l5s1_pos = parameters.iloc[2:frame-2,136:139].to_numpy()
    box_pos = parameters.iloc[2:frame-2,139:142].to_numpy()
    box_acc = parameters.iloc[2:frame-2,142:145].to_numpy()
    inertia = inertia.iloc[:,:].to_numpy()
    landmark = joint.iloc[:frame,1:].to_numpy()
    ha = np.zeros(shape=(frame-4,6))
    
    seg_ma(ha, mass, box_mass, seg_acc, seg)
    seg_acc2 = seg_acc.copy()
    seg_moment_inertia(seg_ang, inertia)

    for i in range(len(seg_com)):
            seg_com[i,::3] -= l5s1_pos[i,0]
            seg_com[i,1::3] -= l5s1_pos[i,1]
            seg_com[i,2::3] -= l5s1_pos[i,2]

    b = bottom_up(mass, frame, FP_data, l5s1_pos, seg_acc, seg_com, seg_ang)
    bottom_up_f.append(b[0])
    bottom_up_m.append(b[1])
    b_compressive_force = compressive_force(frame,b[1])
    bottom_up_c.append(b_compressive_force)
    
    t1 = top_down(mass, frame, seg_acc, FP_data, box_pos, l5s1_pos, seg_com, seg_ang)
    t_compressive_force = compressive_force(frame, t1[1])
    top_down_f.append(t1[0])
    top_down_m.append(t1[1])
    top_down_c.append(t_compressive_force)    

    t2 = top_dwon2(mass, frame, box_mass, seg_acc2, ha, seg_com, seg_ang)
    t2_compressive_force = compressive_force(frame, t2[1])
    top_down_f2.append(t2[0])
    top_down_m2.append(t2[1])
    top_down_c2.append(t2_compressive_force) 
    
    t3 = top_down3(mass, frame, seg_acc, FP_data, box_pos, l5s1_pos, seg_com, seg_ang, box_acc, box_mass)
    t3_compressive_force = compressive_force(frame, t3[1])
    top_down_f3.append(t3[0])
    top_down_m3.append(t3[1])
    top_down_c3.append(t3_compressive_force)
    return b, b_compressive_force, landmark


def save_data(file):
    wb = Workbook()
    ws = wb.active
    ws.title = "L5S1_result"
    for col in range(2,10,1):
        title = ws.cell(1,col-1,title_list1[col-2])
        title.font = Font(bold = True)
        title.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
        title1 = ws.cell(1,col+7,title_list2[col-2])
        title1.font = Font(bold = True)
        title1.fill = PatternFill(start_color="EEAD0E", end_color="EEAD0E", fill_type="solid")
        title2 = ws.cell(1,col+15,title_list3[col-2])
        title2.font = Font(bold = True)
        title2.fill = PatternFill(start_color="EEE9BF", end_color="EEE9BF", fill_type="solid")
        title3 = ws.cell(1,col+23,title_list3[col-2])
        title3.font = Font(bold = True)
        title3.fill = PatternFill(start_color="66CDAA", end_color="66CDAA", fill_type="solid")        
    
    combine_data = np.hstack((bottom_up_f[0],bottom_up_m[0],np.sum(bottom_up_m[0],axis=1, keepdims=True),bottom_up_c[0],top_down_f[0],top_down_m[0],np.sum(top_down_m[0],axis=1, keepdims=True),
                              top_down_c[0],top_down_f2[0],top_down_m2[0],np.sum(top_down_m2[0],axis=1, keepdims=True),top_down_c2[0],top_down_f3[0],top_down_m3[0],np.sum(top_down_m3[0],axis=1, keepdims=True),top_down_c3[0]))
    data_mean = []
    data_max = []
    for i in range(combine_data.shape[1]):        
        data_mean.append(np.mean(abs(combine_data[:,i])))
        data_max.append(np.max(abs(combine_data[:,i])))
    combine_data = np.insert(combine_data,0,data_max,axis=0)
    combine_data = np.insert(combine_data,0,data_mean,axis=0)
    
    combine_data = combine_data.tolist()
    for i in combine_data:
        ws.append(i)   

    for i in range(2,34):
        ws.column_dimensions[get_column_letter(i)].width = 20.5
    ws.insert_cols(1,1)
    ws['A2'] = 'Average'
    ws['A3'] = 'Maximum'    
    wb.save(filename = file)  
    
def save_gif(filename, ani):
    gif_file = os.path.join(os.path.splitext(filename)[0]+ "." + 'gif')
    ani.save(gif_file, writer='pillow')









