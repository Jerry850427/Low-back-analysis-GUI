# -*- coding: utf-8 -*-
"""
Created on Thu Dec  5 17:43:40 2019

@author: 高漢佑
"""
from openpyxl import load_workbook, Workbook
from scipy.interpolate import CubicSpline
from datetime import timedelta
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import scipy.signal as signal
from openpyxl import load_workbook
import sys
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import animation
import mpl_toolkits.mplot3d as plt3d
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d import proj3d
from openpyxl.styles import Font, Color, Fill, PatternFill
from openpyxl.utils import get_column_letter
import os

title_list1 = ['Mx(Top-down)','My(Top-down)','Mz(Top-down)','M_total(Top-down)','CF(Top-down)']              
kinect1 = []
z = [4,10,16,22,28,37,43,52,67,73]
s = [1,26,27,28,29,30,31,47,48,49,59,60,61,65,66,67,68,69,70]
s2 = [25,26,27,28,29,30,46,47,48,58,59,60,64,65,66,67,68,69]
joint = []

def segment_properties(gender):         #get segment propertities based on gender
    if gender == "男性":
        f = open("properties/Male.txt",'r') 
    elif gender == "女性":
        f = open("properties/Female.txt",'r')
    else:
        sys.exit("Error: invalid gender number!")
    return f

def frange(start,stop, step):
    while start <= stop:
        yield round(start,3)
        start +=step     

def time_strip(date,file,d):       
    (h0,m0,s0,ms0) = date[0].split('.')
    time0 = int(h0)*3600000 + int(m0)*60000 + int(s0)*1000 + int(ms0)    
    for i in range(1,len(date)+1):        
        (h,m,s,ms) = date[i-1].split('.')
        time1 = int(h)*3600000 + int(m)*60000 + int(s)*1000 + int(ms)        
        delta_time = round((time1 - time0)*0.001,3)
        file.iloc[:,0][i-1] =  delta_time         
        d.append(delta_time)
    

def cubic_spline(t,x,time_1000hz):    
    fs = CubicSpline(t,x)
    xs = fs(time_1000hz)    
    return xs

def low_pass(data,cutoff,fs,order):    
    nyq = 0.5*fs
    normal_cutoff = cutoff / nyq
    b, a = signal.butter(order, normal_cutoff, output = 'ba')
    j = signal.filtfilt(b,a,data)
    return j

def acc_cal(data):
    vel = []
    acc = []       
    for i in range(1,len(data)-1):        
        vel.append((data[i+1]-data[i-1])/0.016)
    vel = low_pass(vel,8,125,4)
    for i in range(1,len(vel)-1):
        acc.append((vel[i+1]-vel[i-1])/0.016)
    acc = low_pass(acc,8,125,4)
    return acc

def top_down(mass, frame, box_mass, v1, upper_seg_acc):
    k = mass*g 
    upper_seg_mg = np.zeros(shape=(1,3))
    cross_mg = np.zeros(shape=(frame-4,3))
    cross_ma = np.zeros(shape=(frame-4,3))
    for i in range(7,15,1):
        upper_seg_mg +=  seg[i][1]*k
    upper_seg_mg +=  box_mass*g
    for i in range(0,24,3):        
        if i ==6 or i ==9:
            cross_mg = cross_mg + np.cross(v1[:,i:i+3],g*(mass*seg[int(i/3)+7][1]+box_mass/2))
            cross_ma = cross_ma + np.cross(v1[:,i:i+3],(mass*seg[int(i/3)+7][1]+box_mass/2)*upper_seg_acc[:,i:i+3])
        else:
            cross_mg = cross_mg + np.cross(v1[:,i:i+3],k*seg[int(i/3)+7][1])
            cross_ma = cross_ma + np.cross(v1[:,i:i+3],(mass*seg[int(i/3)+7][1])*upper_seg_acc[:,i:i+3])
    top_down_L5S1_moment = -1*(-cross_mg+cross_ma) 
    kinect1.append(top_down_L5S1_moment)
    return top_down_L5S1_moment  

def compressive_force(frame, moment):    
    c = np.zeros(shape=(frame-4,1))     
    for i,j in zip(moment,range(c.shape[0])):        
        cf = 1067.6+1.219*i[0]+0.083*(i[0]**2)-0.0001*(i[0]**3)+3.229*i[1]+0.119*(i[1]**2)-0.0001*(i[1]**3)+0.862*i[2]+0.393*(i[2]**2)-0.0001*(i[2]**3)
        c[j,0] = cf
    kinect1.append(c)
    return c    
    
time = []
time2 = []
time3 = []
dtime = []
dtime2 = []
dtime3 = []
seg = []
g = np.array([0, 0, -9.8])

def get_data(dic, file, time_data):
    global hu
    global data2
    kinect_origin = pd.read_excel(file)
    box_mass = float(dic["box"])
    mass = float(dic["weight"])
    k = segment_properties(dic["sex"])
    
    for i in kinect_origin.columns:
        joint.append(i)
    joint.insert(62,joint.pop(joint.index('Neck_X')))
    joint.insert(63,joint.pop(joint.index('Neck_Y')))
    joint.insert(64,joint.pop(joint.index('Neck_Z')))
    joint.insert(47,joint.pop(joint.index('SpineShoulder_X')))
    joint.insert(48,joint.pop(joint.index('SpineShoulder_Y')))
    joint.insert(49,joint.pop(joint.index('SpineShoulder_Z')))    
    kinect_origin = kinect_origin.loc[:,joint]
    for line in k:
        (key, val) = line.split()
        val = float(val)
        seg.append([key,val])
    
    t = kinect_origin.iloc[:,0].values
    for i in t:
        time.append(i.strip())
    time_strip(time,kinect_origin,dtime)
    time_1000hz = list(frange(dtime[0],dtime[-1],0.001))
    kinect_origin = kinect_origin.drop(['Frame'],axis = 1)
    data = kinect_origin.to_numpy()
    
    interpolate_data = np.zeros(shape=(len(time_1000hz),data.shape[1]))
    interpolate_data[:,0] = time_1000hz
    for i in range(1,data.shape[1]):   
        interpolate_data[:,i] =  cubic_spline(data[:,0], data[:,i], time_1000hz)
    d1 = round(float(time_data[0]),3)
    d2 = round(float(time_data[1]),3)
    m_time = list(frange(d1,d2+0.008,0.008))        # m_time = list(frange(d1,d2+0.008,0.008))

    frame = len(m_time)
    data2 = np.zeros(shape=(frame,data.shape[1]))
    upper_seg_com = np.zeros(shape = (frame,25))
    
    for i in range(len(m_time)):
        data2[i,:] = interpolate_data[int(m_time[i]*1000),:]
        
    upper_seg_com[:,0] = data2[:,0]    
    for i in range(1,data2.shape[1]):
        data2[:,i] = low_pass(data2[:,i],3, 125,4)
    for i in z:
        data2[:,i] = data2[:,i-3]+ 2*(data2[:,55]-data2[:,i-3])
        data2[:,i+1] = data2[:,i-2]
        data2[:,i+2] = data2[:,i-1]
    upper_seg_acc = np.zeros(shape=(frame-4,24))
    upper_seg_com[:,1:4] = data2[:,55:58] + (data2[:,61:64] - data2[:,55:58])*seg[22][1]   #RTA
    upper_seg_com[:,4:7] = data2[:,61:64] + (data2[:,31:34] - data2[:,61:64])*seg[23][1]   #RHE
    upper_seg_com[:,7:10] = data2[:,70:73] + (data2[:,19:22] - data2[:,70:73])*seg[24][1]  #LHA
    upper_seg_com[:,10:13] = data2[:,73:76] + (data2[:,22:25] - data2[:,73:76])*seg[25][1] #RHA
    upper_seg_com[:,13:16] = data2[:,7:10] + (data2[:,70:73] - data2[:,7:10])*seg[26][1]   #LFA
    upper_seg_com[:,16:19] = data2[:,10:13] + (data2[:,73:76] - data2[:,10:13])*seg[27][1] #RFA
    upper_seg_com[:,19:22] = data2[:,49:52] + (data2[:,7:10] - data2[:,49:52])*seg[28][1]  #LAR
    upper_seg_com[:,22:25] = data2[:,52:55] + (data2[:,10:13] - data2[:,52:55])*seg[29][1] #RAR    
    
    v1 = upper_seg_com.copy()
    
    for i in range(len(upper_seg_com)):    
        v1[i,1::3] -= data2[i,55]
        v1[i,2::3] -= data2[i,56]
        v1[i,3::3] -= data2[i,57]
    v1 = np.delete(v1,[0,1,frame-2,frame-1],0)
    v1 = np.delete(v1,0,1)
    
    
    for i in range(1,upper_seg_com.shape[1]):    
        upper_seg_acc[:,i-1] = acc_cal(upper_seg_com[:,i])         
    
    result= top_down(mass, frame, box_mass, v1, upper_seg_acc)
    cf = compressive_force(frame, result)
    hu = np.delete(data2,[25,26,27,28,29,30,46,47,48,58,59,60,64,65,66,67,68,69],1)
    return result, cf, hu

def skeleton(data):  
    hu = data[0]
    joint_plot = [[1,13],[1,34],[28,34],[4,16],[4,37],[31,37],[28,46],[31,46],[46,49],[25,49],[19,52],[7,52],[7,40],[40,49],[22,55],[10,55],[10,43],[43,49]]
    
    

    def update_graph(num):
        num += 2    
        ax.lines = []       
        graph._offsets3d = (hu[:,1::3][num],hu[:,2::3][num],hu[:,3::3][num])
        for i in joint_plot:
            line = plt3d.art3d.Line3D(hu[:,i][num],hu[:,[i[0]+1,i[1]+1]][num],hu[:,[i[0]+2,i[1]+2]][num])
            ax.add_line(line)
        return graph, line
    

    
    class Arrow3D(FancyArrowPatch):
        def __init__(self, xs, ys, zs, *args, **kwargs):
            FancyArrowPatch.__init__(self, (0,0), (0,0), *args, **kwargs)
            self._verts3d = xs, ys, zs
    
        def draw(self, renderer):
            xs3d, ys3d, zs3d = self._verts3d
            xs, ys, zs = proj3d.proj_transform(xs3d, ys3d, zs3d, renderer.M)
            self.set_positions((xs[0],ys[0]),(xs[1],ys[1]))
            FancyArrowPatch.draw(self, renderer)          
    
    fig = plt.figure()
    ax = Axes3D(fig)
    ax.set_xlabel('x')
    ax.set_ylabel('y')
    ax.set_zlabel('z')
    ax.set_xlim(-1, 1)
    ax.set_ylim(-1, 1)
    ax.set_zlim(-1, 2)
    arrow_prop_dict = dict(mutation_scale=20, arrowstyle='->',shrinkA=0, shrinkB=0)
    a = Arrow3D([-0.75,-0.25], [-0.75,-0.75], [-0.75,-0.75], **arrow_prop_dict, color='r')
    ax.add_artist(a)
    a = Arrow3D([-0.75,-0.75], [-0.75,-0.25], [-0.75,-0.75], **arrow_prop_dict, color='b')
    ax.add_artist(a)
    a = Arrow3D([-0.75,-0.75], [-0.75,-0.75], [-0.75,0.25], **arrow_prop_dict, color='g')
    ax.add_artist(a)
    ax.text(0, -1, -1, r'$x$')
    ax.text(-0.75, -0.25, -1, r'$y$')
    ax.text(-0.5, -1, 0, r'$z$')
    graph = ax.scatter([],[],[],color="red")
    
    line = ax.plot([],[],[],color = "blue")
    ax.view_init(elev=45, azim=180)
    ani = animation.FuncAnimation(fig, update_graph,frames = hu.shape[0]-4, blit=False, interval = 10)  
    plt.show(block=False)
    
def save_data(file):
    global hu
    global data2
    wb = Workbook()
    ws = wb.active
    
    ws.title = "L5S1_result"
    ws2 = wb.create_sheet("Joint")
    ws3 = wb.create_sheet("kinect_data")
    for col in range(2,7,1):
        title = ws.cell(1,col-1,title_list1[col-2])
        title.font = Font(bold = True)
        title.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    if len(joint) == 77:
        j = np.delete(joint,s)
        for col in range(2,60,1):
            title2 = ws2.cell(1,col-1,j[col-2])
            title2.font = Font(bold = True)
        joint.pop(1)
    elif len(joint) == 76:
        j = np.delete(joint,s2)
        for col in range(2,60,1):
            title2 = ws2.cell(1,col-1,j[col-2])
            title2.font = Font(bold = True)
    hu = hu.tolist()
    data2 = data2.tolist()   
    ws3.append(joint)
    for row in data2:
        ws3.append(row)        
    for row in hu:
        ws2.append(row)
    combine_data = np.hstack((kinect1[0],np.sum(kinect1[0],axis=1, keepdims=True), kinect1[1]))
    data_mean = []
    data_max = []
    for i in range(combine_data.shape[1]):        
        data_mean.append(np.mean(abs(combine_data[:,i])))
        data_max.append(np.max(abs(combine_data[:,i])))
    combine_data = np.insert(combine_data,0,data_max,axis=0)
    combine_data = np.insert(combine_data,0,data_mean,axis=0)
    
    combine_data = combine_data.tolist()
    for i in combine_data:
        ws.append(i)   

    for i in range(2,7):
        ws.column_dimensions[get_column_letter(i)].width = 20.5
    ws.insert_cols(1,1)
    ws['A2'] = 'Average'
    ws['A3'] = 'Maximum'    
    wb.save(filename = file)        

def save_gif(filename, ani):
    gif_file = os.path.join(os.path.splitext(filename)[0]+ "." + 'gif')
    ani.save(gif_file, writer='pillow')


def get_data2(dic, file, time_data):
    global hu
    global data2
    k = file.split("#")
    
    kinect_origin = pd.read_excel(k[0])
    kinect_origin2 = pd.read_excel(k[1])
    
    for i in kinect_origin.columns:
        joint.append(i)
    joint.insert(62,joint.pop(joint.index('Neck_X')))
    joint.insert(63,joint.pop(joint.index('Neck_Y')))
    joint.insert(64,joint.pop(joint.index('Neck_Z')))
    joint.insert(47,joint.pop(joint.index('SpineShoulder_X')))
    joint.insert(48,joint.pop(joint.index('SpineShoulder_Y')))
    joint.insert(49,joint.pop(joint.index('SpineShoulder_Z')))    
    kinect_origin = kinect_origin.loc[:,joint]
    
    box_mass = float(dic["box"])
    mass = float(dic["weight"])
    k = segment_properties(dic["sex"])
    
    for line in k:
        (key, val) = line.split()
        val = float(val)
        seg.append([key,val])
        
    t = kinect_origin.iloc[:,0].values
    t2 = kinect_origin2.iloc[:,0].values
    for i in t:
        time.append(i.strip())
    for j in t2:
        time2.append(j.strip())
    time_strip(time,kinect_origin,dtime)
    time_strip(time2,kinect_origin2,dtime2)
    
    time_1000hz = list(frange(dtime[0],dtime[-1],0.001))
    time_1000hz2 = list(frange(dtime2[0],dtime2[-1],0.001))
    kinect_origin = kinect_origin.drop(['Frame'],axis = 1)
    kinect_origin2 = kinect_origin2.drop(['Frame'],axis = 1)
    
    data = kinect_origin.to_numpy()
    data_180 = kinect_origin2.to_numpy()
    
    interpolate_data = np.zeros(shape=(len(time_1000hz),data.shape[1]))
    interpolate_data[:,0] = time_1000hz
    
    interpolate_data2 = np.zeros(shape=(len(time_1000hz2),data_180.shape[1]))
    interpolate_data2[:,0] = time_1000hz2
    
    for i in range(1,data.shape[1]):   
        interpolate_data[:,i] =  cubic_spline(data[:,0], data[:,i], time_1000hz)
    for i in range(1,data_180.shape[1]): 
        interpolate_data2[:,i] =  cubic_spline(data_180[:,0], data_180[:,i], time_1000hz2)    
        
    d1 = round(float(time_data[0].split("#")[0]),3)
    d2 = round(float(time_data[1].split("#")[0]),3)
    t1 = round(float(time_data[0].split("#")[1]),3)
    t2 = round(float(time_data[1].split("#")[1]),3)

    m_time = list(frange(d1,d2+0.008,0.008))
    m_time2 = list(frange(t1,t2+0.008,0.008))

    frame = len(m_time)
    frame2 = len(m_time2)
    data2 = np.zeros(shape=(frame,data.shape[1]))
    data2_180 = np.zeros(shape=(frame2,data_180.shape[1]))
    
    upper_seg_com = np.zeros(shape = (frame,25))
    upper_seg_com_2 = np.zeros(shape = (frame2,25))
    
    for i in range(len(m_time)):
        data2[i,:] = interpolate_data[int(m_time[i]*1000),:]
    for i in range(len(m_time2)):
        data2_180[i,:] = interpolate_data2[int(m_time2[i]*1000),:]
        
    upper_seg_com[:,0] = data2[:,0]    
    for i in range(1,data2.shape[1]):
        data2[:,i] = low_pass(data2[:,i], 5, 125, 4)
    for i in range(1,data2_180.shape[1]):
        data2_180[:,i] = low_pass(data2_180[:,i], 5, 125, 4)
    
    data2[:,55] = data2_180[:,34]
    data2[:,56] = data2_180[:,35]
    data2[:,57] = data2_180[:,36]
    
    for i in z:
        data2[:,i] = data2[:,i-3]+ 2*(data2[:,55]-data2[:,i-3])
        data2[:,i+1] = data2[:,i-2]
        data2[:,i+2] = data2[:,i-1]
    print(seg[22][1])    
    upper_seg_acc = np.zeros(shape=(frame-4,24))
    upper_seg_com[:,1:4] = data2[:,55:58] + (data2[:,61:64] - data2[:,55:58])*seg[22][1]   #RTA
    upper_seg_com[:,4:7] = data2[:,61:64]    #RHE
    upper_seg_com[:,7:10] = data2[:,70:73] + (data2[:,19:22] - data2[:,70:73])*seg[24][1]  #LHA
    upper_seg_com[:,10:13] = data2[:,73:76] + (data2[:,22:25] - data2[:,73:76])*seg[25][1] #RHA
    upper_seg_com[:,13:16] = data2[:,7:10] + (data2[:,70:73] - data2[:,7:10])*seg[26][1]   #LFA
    upper_seg_com[:,16:19] = data2[:,10:13] + (data2[:,73:76] - data2[:,10:13])*seg[27][1] #RFA
    upper_seg_com[:,19:22] = data2[:,49:52] + (data2[:,7:10] - data2[:,49:52])*seg[28][1]  #LAR
    upper_seg_com[:,22:25] = data2[:,52:55] + (data2[:,10:13] - data2[:,52:55])*seg[29][1] #RAR

    
    v1 = upper_seg_com.copy()
    
    for i in range(len(upper_seg_com)):    
        v1[i,1::3] -= data2[i,55]
        v1[i,2::3] -= data2[i,56]
        v1[i,3::3] -= data2[i,57]
        
    v1 = np.delete(v1,[0,1,frame-2,frame-1],0)
    v1 = np.delete(v1,0,1)
    
    for i in range(1,upper_seg_com.shape[1]):    
        upper_seg_acc[:,i-1] = acc_cal(upper_seg_com[:,i])         
    
    result= top_down(mass, frame, box_mass, v1, upper_seg_acc)
    cf = compressive_force(frame, result)
    hu = np.delete(data2,[25,26,27,28,29,30,46,47,48,58,59,60,64,65,66,67,68,69],1)
    return result, cf, hu

def get_data3(dic, file, time_data):
    global hu
    global data2
       
    kinect_origin = pd.read_excel(file,sheet_name = 'kinect_data')
    for i in kinect_origin.columns:
        joint.append(i)
    kinect_origin = kinect_origin.loc[:,joint]
      
    
    box_mass = float(dic["box"])
    mass = float(dic["weight"])
    k = segment_properties(dic["sex"])
    
    for line in k:
        (key, val) = line.split()
        val = float(val)
        seg.append([key,val]) 
        
    data = kinect_origin.to_numpy()
    
    for i in z:
        data[:,i] = data[:,i-3]+ 2*(data[:,55]-data[:,i-3])
        data[:,i+1] = data[:,i-2]
        data[:,i+2] = data[:,i-1]       

    
    frame = len(data[:,0])    
    upper_seg_acc = np.zeros(shape=(frame-4,24))
    upper_seg_com = np.zeros(shape = (frame,25))
    upper_seg_com[:,0] = data[:,0] 
    upper_seg_com[:,1:4] = data[:,55:58] + (data[:,61:64] - data[:,55:58])*seg[22][1]   #RTA
    upper_seg_com[:,4:7] = data[:,61:64] + (data[:,31:34] - data[:,61:64])*seg[23][1]   #RHE
    upper_seg_com[:,7:10] = data[:,70:73] + (data[:,19:22] - data[:,70:73])*seg[24][1]  #LHA
    upper_seg_com[:,10:13] = data[:,73:76] + (data[:,22:25] - data[:,73:76])*seg[25][1] #RHA
    upper_seg_com[:,13:16] = data[:,7:10] + (data[:,70:73] - data[:,7:10])*seg[26][1]   #LFA
    upper_seg_com[:,16:19] = data[:,10:13] + (data[:,73:76] - data[:,10:13])*seg[27][1] #RFA
    upper_seg_com[:,19:22] = data[:,49:52] + (data[:,7:10] - data[:,49:52])*seg[28][1]  #LAR
    upper_seg_com[:,22:25] = data[:,52:55] + (data[:,10:13] - data[:,52:55])*seg[29][1] #RAR
    
    v1 = upper_seg_com.copy()

    for i in range(len(upper_seg_com)): 
        v1[i,1::3] -= data[i,55]
        v1[i,2::3] -= data[i,56]
        v1[i,3::3] -= data[i,57]
    
    v1 = np.delete(v1,[0,1,frame-2,frame-1],0)
    v1 = np.delete(v1,0,1)    
    for i in range(1,upper_seg_com.shape[1]): 
        upper_seg_acc[:,i-1] = acc_cal(upper_seg_com[:,i]) 
    data2 = data.copy()
    result= top_down(mass, frame, box_mass, v1, upper_seg_acc)
    cf = compressive_force(frame, result)
    hu = np.delete(data2,[25,26,27,28,29,30,46,47,48,58,59,60,64,65,66,67,68,69],1)
    return result, cf, hu